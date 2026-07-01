"""
AttendanceAI - Utility Helpers
"""
import os
import io
import csv
import json
import secrets
import logging
from datetime import datetime, timedelta
from functools import wraps
from flask import request, current_app, abort
from flask_login import current_user
from flask_mail import Message

logger = logging.getLogger(__name__)


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                from flask import redirect, url_for
                return redirect(url_for('auth.login'))
            if current_user.role.value not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def generate_token(length: int = 32) -> str:
    return secrets.token_urlsafe(length)


def generate_reset_token_expiry() -> datetime:
    return datetime.utcnow() + timedelta(hours=1)


def send_verification_email(mail, user, token: str):
    from flask import url_for, render_template
    try:
        verify_url = url_for('auth.verify_email', token=token, _external=True)
        msg = Message(subject='Verify Your Email - AttendanceAI', recipients=[user.email],
                      html=render_template('email/verify_email.html', user=user, verify_url=verify_url))
        mail.send(msg)
        logger.info(f"Verification email sent to {user.email}")
    except Exception as e:
        logger.error(f"Failed to send verification email: {e}")


def send_password_reset_email(mail, user, token: str):
    from flask import url_for, render_template
    try:
        reset_url = url_for('auth.reset_password', token=token, _external=True)
        msg = Message(subject='Reset Your Password - AttendanceAI', recipients=[user.email],
                      html=render_template('email/reset_password.html', user=user, reset_url=reset_url))
        mail.send(msg)
        logger.info(f"Password reset email sent to {user.email}")
    except Exception as e:
        logger.error(f"Failed to send password reset email: {e}")


def send_attendance_notification(mail, student_user, attendance_record):
    from flask import render_template
    try:
        msg = Message(subject='Attendance Marked - AttendanceAI', recipients=[student_user.email],
                      html=render_template('email/attendance_notification.html',
                                           user=student_user, record=attendance_record))
        mail.send(msg)
    except Exception as e:
        logger.error(f"Failed to send attendance notification: {e}")


def send_low_attendance_alert(mail, student_user, subject, percentage):
    from flask import render_template
    try:
        msg = Message(subject=f'Low Attendance Alert: {subject} - AttendanceAI', recipients=[student_user.email],
                      html=render_template('email/low_attendance.html',
                                           user=student_user, subject=subject, percentage=percentage))
        mail.send(msg)
    except Exception as e:
        logger.error(f"Failed to send low attendance alert: {e}")


def log_action(db, action: str, resource: str = None, resource_id: int = None,
               details: dict = None, user_id: int = None):
    from app.models import AuditLog
    try:
        log = AuditLog(
            user_id=user_id or (current_user.id if current_user.is_authenticated else None),
            action=action, resource=resource, resource_id=resource_id,
            details=json.dumps(details) if details else None,
            ip_address=request.remote_addr if request else None,
            user_agent=request.headers.get('User-Agent', '')[:255] if request else None
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logger.error(f"Failed to create audit log: {e}")


def create_notification(db, user_id: int, title: str, message: str,
                         notif_type: str = 'system', link: str = None):
    from app.models import Notification, NotificationType
    try:
        notif = Notification(user_id=user_id, title=title, message=message,
                             type=NotificationType[notif_type.upper()], link=link)
        db.session.add(notif)
        db.session.commit()
    except Exception as e:
        logger.error(f"Failed to create notification: {e}")


def export_attendance_csv(records) -> io.BytesIO:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Roll Number', 'Student Name', 'Class', 'Status', 'Time In', 'Confidence', 'Marked By'])
    for record in records:
        writer.writerow([
            record.date, record.student.roll_number, record.student.user.full_name,
            record.class_.name, record.status.value,
            record.time_in.strftime('%H:%M:%S') if record.time_in else '',
            f"{record.confidence_score:.2f}" if record.confidence_score else '',
            record.marked_by
        ])
    output.seek(0)
    return io.BytesIO(output.getvalue().encode())


def export_attendance_excel(records) -> io.BytesIO:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        headers = ['Date', 'Roll Number', 'Student Name', 'Class', 'Status', 'Time In', 'Confidence Score', 'Marked By']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        status_fills = {
            'present': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            'absent': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            'late': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'excused': PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        }

        for row, record in enumerate(records, 2):
            data = [
                record.date.strftime('%Y-%m-%d'), record.student.roll_number,
                record.student.user.full_name, record.class_.name,
                record.status.value.title(),
                record.time_in.strftime('%H:%M:%S') if record.time_in else 'N/A',
                f"{record.confidence_score:.2f}" if record.confidence_score else 'N/A',
                record.marked_by.replace('_', ' ').title()
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 5:
                    fill = status_fills.get(record.status.value)
                    if fill:
                        cell.fill = fill

        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return export_attendance_csv(records)


def export_attendance_pdf(records, title: str = "Attendance Report") -> io.BytesIO:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18,
                                     textColor=colors.HexColor('#2563EB'))
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        table_data = [['Date', 'Roll No.', 'Student Name', 'Class', 'Status', 'Time In', 'Confidence']]
        for record in records:
            table_data.append([
                record.date.strftime('%Y-%m-%d'), record.student.roll_number,
                record.student.user.full_name, record.class_.name,
                record.status.value.title(),
                record.time_in.strftime('%H:%M') if record.time_in else 'N/A',
                f"{record.confidence_score:.0%}" if record.confidence_score else 'N/A'
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return io.BytesIO(b"PDF generation failed")


def calculate_attendance_stats(records) -> dict:
    total = len(records)
    if total == 0:
        return {'total': 0, 'present': 0, 'absent': 0, 'late': 0, 'excused': 0, 'percentage': 0.0}
"""
AttendanceAI - Utility Helpers
"""
import os
import io
import csv
import json
import secrets
import logging
from datetime import datetime, timedelta
from functools import wraps
from flask import request, current_app, abort
from flask_login import current_user
from flask_mail import Message

logger = logging.getLogger(__name__)


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                from flask import redirect, url_for
                return redirect(url_for('auth.login'))
            if current_user.role.value not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def generate_token(length: int = 32) -> str:
    return secrets.token_urlsafe(length)


def generate_reset_token_expiry() -> datetime:
    return datetime.utcnow() + timedelta(hours=1)


def send_verification_email(mail, user, token: str):
    from flask import url_for, render_template
    try:
        verify_url = url_for('auth.verify_email', token=token, _external=True)
        msg = Message(subject='Verify Your Email - AttendanceAI', recipients=[user.email],
                      html=render_template('email/verify_email.html', user=user, verify_url=verify_url))
        mail.send(msg)
        logger.info(f"Verification email sent to {user.email}")
    except Exception as e:
        logger.error(f"Failed to send verification email: {e}")


def send_password_reset_email(mail, user, token: str):
    from flask import url_for, render_template
    try:
        reset_url = url_for('auth.reset_password', token=token, _external=True)
        msg = Message(subject='Reset Your Password - AttendanceAI', recipients=[user.email],
                      html=render_template('email/reset_password.html', user=user, reset_url=reset_url))
        mail.send(msg)
        logger.info(f"Password reset email sent to {user.email}")
    except Exception as e:
        logger.error(f"Failed to send password reset email: {e}")


def send_attendance_notification(mail, student_user, attendance_record):
    from flask import render_template
    try:
        msg = Message(subject='Attendance Marked - AttendanceAI', recipients=[student_user.email],
                      html=render_template('email/attendance_notification.html',
                                           user=student_user, record=attendance_record))
        mail.send(msg)
    except Exception as e:
        logger.error(f"Failed to send attendance notification: {e}")


def send_low_attendance_alert(mail, student_user, subject, percentage):
    from flask import render_template
    try:
        msg = Message(subject=f'Low Attendance Alert: {subject} - AttendanceAI', recipients=[student_user.email],
                      html=render_template('email/low_attendance.html',
                                           user=student_user, subject=subject, percentage=percentage))
        mail.send(msg)
    except Exception as e:
        logger.error(f"Failed to send low attendance alert: {e}")


def log_action(db, action: str, resource: str = None, resource_id: int = None,
               details: dict = None, user_id: int = None):
    from app.models import AuditLog
    try:
        log = AuditLog(
            user_id=user_id or (current_user.id if current_user.is_authenticated else None),
            action=action, resource=resource, resource_id=resource_id,
            details=json.dumps(details) if details else None,
            ip_address=request.remote_addr if request else None,
            user_agent=request.headers.get('User-Agent', '')[:255] if request else None
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logger.error(f"Failed to create audit log: {e}")


def create_notification(db, user_id: int, title: str, message: str,
                         notif_type: str = 'system', link: str = None):
    from app.models import Notification, NotificationType
    try:
        notif = Notification(user_id=user_id, title=title, message=message,
                             type=NotificationType[notif_type.upper()], link=link)
        db.session.add(notif)
        db.session.commit()
    except Exception as e:
        logger.error(f"Failed to create notification: {e}")


def export_attendance_csv(records) -> io.BytesIO:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Roll Number', 'Student Name', 'Class', 'Status', 'Time In', 'Confidence', 'Marked By'])
    for record in records:
        writer.writerow([
            record.date, record.student.roll_number, record.student.user.full_name,
            record.class_.name, record.status.value,
            record.time_in.strftime('%H:%M:%S') if record.time_in else '',
            f"{record.confidence_score:.2f}" if record.confidence_score else '',
            record.marked_by
        ])
    output.seek(0)
    return io.BytesIO(output.getvalue().encode())


def export_attendance_excel(records) -> io.BytesIO:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        headers = ['Date', 'Roll Number', 'Student Name', 'Class', 'Status', 'Time In', 'Confidence Score', 'Marked By']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        status_fills = {
            'present': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            'absent': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            'late': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'excused': PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        }

        for row, record in enumerate(records, 2):
            data = [
                record.date.strftime('%Y-%m-%d'), record.student.roll_number,
                record.student.user.full_name, record.class_.name,
                record.status.value.title(),
                record.time_in.strftime('%H:%M:%S') if record.time_in else 'N/A',
                f"{record.confidence_score:.2f}" if record.confidence_score else 'N/A',
                record.marked_by.replace('_', ' ').title()
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 5:
                    fill = status_fills.get(record.status.value)
                    if fill:
                        cell.fill = fill

        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return export_attendance_csv(records)


def export_attendance_pdf(records, title: str = "Attendance Report") -> io.BytesIO:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=18,
                                     textColor=colors.HexColor('#2563EB'))
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        table_data = [['Date', 'Roll No.', 'Student Name', 'Class', 'Status', 'Time In', 'Confidence']]
        for record in records:
            table_data.append([
                record.date.strftime('%Y-%m-%d'), record.student.roll_number,
                record.student.user.full_name, record.class_.name,
                record.status.value.title(),
                record.time_in.strftime('%H:%M') if record.time_in else 'N/A',
                f"{record.confidence_score:.0%}" if record.confidence_score else 'N/A'
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return io.BytesIO(b"PDF generation failed")


def calculate_attendance_stats(records) -> dict:
    total = len(records)
    if total == 0:
        return {'total': 0, 'present': 0, 'absent': 0, 'late': 0, 'excused': 0, 'percentage': 0.0}

    from app.models import AttendanceStatus
    present = sum(1 for r in records if r.status == AttendanceStatus.PRESENT)
    absent = sum(1 for r in records if r.status == AttendanceStatus.ABSENT)
    late = sum(1 for r in records if r.status == AttendanceStatus.LATE)
    excused = sum(1 for r in records if r.status == AttendanceStatus.EXCUSED)
    attended = present + late
    percentage = (attended / total) * 100 if total > 0 else 0

    return {'total': total, 'present': present, 'absent': absent, 'late': late,
            'excused': excused, 'attended': attended, 'percentage': round(percentage, 1)}


def paginate_query(query, page: int, per_page: int = 20):
    return query.paginate(page=page, per_page=per_page, error_out=False)